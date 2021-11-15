from openpyxl import load_workbook
import datetime
import logging

logging.basicConfig(level=logging.INFO, filename="assignment_log.log", filemode='a', format='%(process)d - [%(levelname)s ] - %(message)s')

logging.info('Starting weekend assignment program')

# note that january Summary only March 2017 to February 2018, VOC only April 2016 to February 2018
# note that march Summary only April 2017 to March 2018, VOC only April 2016 to March 2018

## Program
# get user input
print('Enter month file to use')
month_file = input()
print('Enter month to check')
given_month = input()
print('Enter year to check')
given_year = input()
derived_filename = f"expedia_report_monthly_{month_file}_2018.xlsx"
logging.info(f'opening {derived_filename} to check for {given_month} {given_year}')
try:
    # open file and begin
    derived_workbook = load_workbook(derived_filename)
    logging.info(f'getting {derived_filename} sheets')
    try:
        summary_sheet = derived_workbook["Summary Rolling MoM"]
        voc_sheet = derived_workbook["VOC Rolling MoM"]
    except Exception as e:
        logging.error(f'error getting {derived_filename} sheets', exc_info=True)

    # summary data dictionary
    summary_data = {}

    logging.info(f'extracting {summary_sheet} sheet data')
    try:
        # get data from cells
        for row in summary_sheet.iter_rows(min_row=2, max_row=13, min_col=1, max_col = 6, values_only=True):
            # ensure keys are uniform -> Month Year format
            if type(row[0]) == datetime.datetime:
                temp_obj = datetime.datetime.strptime(row[0].date().isoformat()[5:7],'%m')
                data_id = temp_obj.strftime('%B') + " " + row[0].date().isoformat()[0:4]
            else:
                data_id = row[0] + " 2018"
            # get values
            datum = {
                "Calls Offered" : row[1],
                "Abandon after 30s" : row[2],
                "FCR" : row[3],
                "DSAT" : row[4],
                "CSAT" : row[5]
            }
            # store values with key in dictionary
            summary_data[data_id] = datum
        logging.info(f'{summary_sheet} data extraction completed')

        # bool variable to check if data is found
        found_data = False
        # print all values of desired month
        print(f'Getting information about {given_month} {given_year} Summary')
        logging.info(f'searching extracted summary data for {given_month} {given_year}')
        for data_id, data_values in summary_data.items():
            # check if key contains short month name and year
            if (given_month[:3] in data_id.lower()) and (given_year in data_id):
                logging.info(f'found {given_month} {given_year} with values {data_values.items()}')
                found_data = True
                # get individual values
                for key in data_values:
                    if key == "Calls Offered":
                        print(key + ':', data_values[key])
                    else:
                        print(f'{key} : {(data_values[key]*100):.2f}%')
        if not found_data:
            print(f'Summary data for {given_month} {given_year} not found')
            logging.info(f'no summary data for {given_month} {given_year} found')
        print('')
    except Exception as e:
        logging.error(f'error extracting data from {summary_sheet}', exc_info=True)
    
    # VOC data dictionary
    voc_data = {}

    logging.info(f'extracing {voc_sheet} sheet data')
    try:
        for column in voc_sheet.iter_cols(min_row=1, max_row=19, min_col=2, max_col=25, values_only=True):
            # ensure keys are uniform -> Month Year format
            if type(column[0]) == datetime.datetime:
                temp_obj = datetime.datetime.strptime(column[0].date().isoformat()[5:7],'%m')
                data_id = temp_obj.strftime('%B') + " " + column[0].date().isoformat()[0:4]
            elif type(column[0]) == str:
                data_id = column[0] + " 2018"
            else:
                data_id = column[0]
            # get values
            datum = {
                "Base Size" : column[2],
                "Promoters Total" : column[3],
                "Promoters Percent" : column[4],
                "Passives Total" : column[5],
                "Passives Percent" : column[6],
                "Detractors Total" : column[7],
                "Detractors Percent" : column[8],
                "Overall NPS Percent" : column[12],
                "Sat with Agent Percent" : column[15],
                "DSat with Agent Percent" : column[18]
            }
            # store values with key in dictionary
            voc_data[data_id] = datum
        logging.info(f'{voc_sheet} data extraction completed')

        # print all values of desired month
        print(f'Getting information about {given_month} {given_year} VOC')
        logging.info(f'searching extracted VOC data for {given_month} {given_year}')
        # bool variable to check if data is found
        found_data = False
        # temp string var for net promoter score
        net_promoter_score = ""
        for data_id, data_values in voc_data.items():
            # check if key contains short month name and year
            if (given_month[:3] in data_id.lower()) and (given_year in data_id):
                logging.info(f'found {given_month} {given_year} with values {data_values.items()}')
                found_data = True
                # get individual values
                for key in data_values:
                    if "Percent" in key:
                        print(f'{key} : {(data_values[key]*100):.2f}%')
                    else:
                        print(key + ':', data_values[key])
                    # check net promoter score
                    if "Total" in key:
                        if "Promoters" in key:
                            if data_values[key] > 200:
                                print('good promoters')
                                net_promoter_score += 'good promoters, '
                            else:
                                print('bad promoters')
                                net_promoter_score += 'bad promoters, '
                        if "Passives" in key:
                            if data_values[key] > 100:
                                print('good passives')
                                net_promoter_score += 'good passives, '
                            else:
                                print('bad passives')
                                net_promoter_score += 'bad passives, '
                        if "Detractors" in key:
                            if data_values[key] > 100:
                                print('good detractors')
                                net_promoter_score += 'good detractors'
                            else:
                                print('bad detractors')
                                net_promoter_score += 'bad detractors'
        if not found_data:
            print(f'VOC data for {given_month} {given_year} not found')
            logging.info(f'no VOC data for {given_month} {given_year} found')
        else:
            logging.info(f'{given_month} {given_year} had {net_promoter_score}')
        print('')
    except Exception as e:
        logging.error(f'error extracting data from {voc_sheet}', exc_info=True)
except Exception as e:
    logging.error(f'error opening {derived_filename}', exc_info=True)
logging.info('Ending weekend assignment program')